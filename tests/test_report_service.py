"""
SentinelOps — Report Service Unit Tests
==========================================
Tests for ``ReportService``, ``ChartBuilder``, and report generation
in CSV, Excel, and PDF formats.

Each test creates isolated services backed by ``tmp_path``.
"""

from __future__ import annotations

import csv
from datetime import datetime, timezone
from pathlib import Path

import pytest

from app.models.alert import AlertCreate, AlertType, Severity
from app.models.report import ReportFormat, ReportRequest
from app.services.alert_service import AlertService
from app.services.analytics_service import AnalyticsService
from app.services.chart_builder import ChartBuilder
from app.services.report_service import ReportNotFoundError, ReportService


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _make_services(tmp_path):
    """Return (AlertService, AnalyticsService, ReportService) triple."""
    alert_svc = AlertService(alerts_dir=tmp_path / "alerts")
    analytics_svc = AnalyticsService(alert_svc)
    report_svc = ReportService(
        alert_svc, analytics_svc, reports_dir=tmp_path / "reports"
    )
    return alert_svc, analytics_svc, report_svc


def _seed_alert(
    svc: AlertService,
    camera_id: str = "cam_01",
    alert_type: AlertType = AlertType.NO_HELMET,
    severity: Severity = Severity.HIGH,
    confidence: float = 0.9,
    image_path: str | None = None,
) -> None:
    svc.create(
        AlertCreate(
            camera_id=camera_id,
            alert_type=alert_type,
            severity=severity,
            confidence=confidence,
            image_path=image_path,
        )
    )


def _seed_mixed_alerts(alert_svc: AlertService, n: int = 6) -> None:
    """Seed a mix of alert types for realistic reports."""
    types = [
        AlertType.NO_HELMET,
        AlertType.NO_VEST,
        AlertType.LOITERING,
        AlertType.PERSON_DETECTED,
        AlertType.RESTRICTED_AREA_ENTRY,
        AlertType.CROWD_FORMATION,
    ]
    cameras = ["cam_01", "cam_02", "cam_03"]
    for i in range(n):
        _seed_alert(
            alert_svc,
            camera_id=cameras[i % len(cameras)],
            alert_type=types[i % len(types)],
            severity=[Severity.LOW, Severity.MEDIUM, Severity.HIGH, Severity.CRITICAL][i % 4],
        )


# ===========================================================================
# ChartBuilder Tests
# ===========================================================================
class TestChartBuilder:
    def test_violations_per_day_chart(self, tmp_path):
        alert_svc, analytics_svc, _ = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        builder = ChartBuilder(output_dir=tmp_path / "charts")
        data = analytics_svc.violations_per_day()
        path = builder.violations_per_day_chart(data)

        assert path.exists()
        assert path.suffix == ".png"
        assert path.stat().st_size > 0

    def test_violations_per_camera_chart(self, tmp_path):
        alert_svc, analytics_svc, _ = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        builder = ChartBuilder(output_dir=tmp_path / "charts")
        data = analytics_svc.violations_per_camera()
        path = builder.violations_per_camera_chart(data)

        assert path.exists()
        assert path.suffix == ".png"

    def test_compliance_pie_chart(self, tmp_path):
        alert_svc, analytics_svc, _ = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        builder = ChartBuilder(output_dir=tmp_path / "charts")
        data = analytics_svc.compliance_rate()
        path = builder.compliance_pie_chart(data)

        assert path.exists()
        assert path.suffix == ".png"

    def test_hourly_trends_chart(self, tmp_path):
        alert_svc, analytics_svc, _ = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        builder = ChartBuilder(output_dir=tmp_path / "charts")
        data = analytics_svc.hourly_trends()
        path = builder.hourly_trends_chart(data)

        assert path.exists()
        assert path.suffix == ".png"

    def test_top_violations_chart(self, tmp_path):
        alert_svc, analytics_svc, _ = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        builder = ChartBuilder(output_dir=tmp_path / "charts")
        data = analytics_svc.top_violation_types()
        path = builder.top_violations_chart(data)

        assert path.exists()
        assert path.suffix == ".png"

    def test_empty_data_no_crash(self, tmp_path):
        """Charts with empty data should not raise."""
        _, analytics_svc, _ = _make_services(tmp_path)
        builder = ChartBuilder(output_dir=tmp_path / "charts")

        assert builder.violations_per_day_chart(analytics_svc.violations_per_day()).exists()
        assert builder.violations_per_camera_chart(analytics_svc.violations_per_camera()).exists()
        assert builder.compliance_pie_chart(analytics_svc.compliance_rate()).exists()
        assert builder.hourly_trends_chart(analytics_svc.hourly_trends()).exists()
        assert builder.top_violations_chart(analytics_svc.top_violation_types()).exists()


# ===========================================================================
# CSV Report Tests
# ===========================================================================
class TestCSVReport:
    def test_file_created(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.CSV))
        path = tmp_path / "reports" / meta.file_path

        assert path.exists()
        assert path.suffix == ".csv"
        assert meta.file_size_bytes > 0

    def test_correct_headers(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_alert(alert_svc)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.CSV))
        path = tmp_path / "reports" / meta.file_path

        with open(path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader)

        assert "alert_id" in headers
        assert "timestamp" in headers
        assert "camera_id" in headers
        assert "alert_type" in headers

    def test_correct_row_count(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc, n=5)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.CSV))
        path = tmp_path / "reports" / meta.file_path

        with open(path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        # 1 header + 5 data rows before the summary section
        data_rows = rows[1:6]
        assert len(data_rows) == 5

    def test_summary_stats_appended(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.CSV))
        path = tmp_path / "reports" / meta.file_path

        content = path.read_text(encoding="utf-8")
        assert "SUMMARY STATISTICS" in content
        assert "Total Violations" in content
        assert "PPE Compliance Rate" in content

    def test_empty_data(self, tmp_path):
        _, _, report_svc = _make_services(tmp_path)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.CSV))
        path = tmp_path / "reports" / meta.file_path

        with open(path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        # Should still have header + summary
        assert len(rows) > 1
        content = path.read_text(encoding="utf-8")
        assert "alert_id" in content  # header present


# ===========================================================================
# Excel Report Tests
# ===========================================================================
class TestExcelReport:
    def test_file_created(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.EXCEL))
        path = tmp_path / "reports" / meta.file_path

        assert path.exists()
        assert path.suffix == ".xlsx"
        assert meta.file_size_bytes > 0

    def test_correct_sheet_names(self, tmp_path):
        from openpyxl import load_workbook

        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        meta = report_svc.generate(
            ReportRequest(format=ReportFormat.EXCEL, include_charts=True)
        )
        path = tmp_path / "reports" / meta.file_path

        wb = load_workbook(path)
        assert "Violations" in wb.sheetnames
        assert "Statistics" in wb.sheetnames
        assert "Charts" in wb.sheetnames

    def test_violations_row_count(self, tmp_path):
        from openpyxl import load_workbook

        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc, n=5)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.EXCEL))
        path = tmp_path / "reports" / meta.file_path

        wb = load_workbook(path)
        ws = wb["Violations"]
        # 1 header + 5 data rows
        assert ws.max_row == 6

    def test_statistics_sheet_content(self, tmp_path):
        from openpyxl import load_workbook

        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.EXCEL))
        path = tmp_path / "reports" / meta.file_path

        wb = load_workbook(path)
        ws = wb["Statistics"]
        assert ws.cell(row=1, column=1).value == "SUMMARY STATISTICS"

    def test_no_charts_sheet_when_disabled(self, tmp_path):
        from openpyxl import load_workbook

        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_alert(alert_svc)

        meta = report_svc.generate(
            ReportRequest(format=ReportFormat.EXCEL, include_charts=False)
        )
        path = tmp_path / "reports" / meta.file_path

        wb = load_workbook(path)
        assert "Charts" not in wb.sheetnames

    def test_charts_sheet_has_images(self, tmp_path):
        from openpyxl import load_workbook

        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        meta = report_svc.generate(
            ReportRequest(format=ReportFormat.EXCEL, include_charts=True)
        )
        path = tmp_path / "reports" / meta.file_path

        wb = load_workbook(path)
        ws = wb["Charts"]
        assert len(ws._images) == 5


# ===========================================================================
# PDF Report Tests
# ===========================================================================
class TestPDFReport:
    def test_file_created(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.PDF))
        path = tmp_path / "reports" / meta.file_path

        assert path.exists()
        assert path.suffix == ".pdf"
        assert meta.file_size_bytes > 0

    def test_valid_pdf(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.PDF))
        path = tmp_path / "reports" / meta.file_path

        # Valid PDF starts with %PDF
        with open(path, "rb") as f:
            header = f.read(5)
        assert header == b"%PDF-"

    def test_reasonable_size(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc, n=10)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.PDF))
        # A PDF with charts and data should be at least a few KB
        assert meta.file_size_bytes > 1000

    def test_without_charts(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_mixed_alerts(alert_svc)

        meta_no_charts = report_svc.generate(
            ReportRequest(format=ReportFormat.PDF, include_charts=False)
        )
        meta_with_charts = report_svc.generate(
            ReportRequest(format=ReportFormat.PDF, include_charts=True)
        )
        # With charts should be larger
        assert meta_with_charts.file_size_bytes > meta_no_charts.file_size_bytes

    def test_custom_title(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_alert(alert_svc)

        meta = report_svc.generate(
            ReportRequest(format=ReportFormat.PDF, title="Custom Test Report")
        )
        # Just verify it generates without error
        path = tmp_path / "reports" / meta.file_path
        assert path.exists()

    def test_with_screenshot(self, tmp_path):
        """PDF embeds screenshots when image_path files exist."""
        from PIL import Image as PILImage

        alert_svc, _, report_svc = _make_services(tmp_path)

        # Create a proper test image using Pillow
        img_dir = tmp_path / "screenshots"
        img_dir.mkdir()
        fake_img = img_dir / "test_screenshot.png"
        img = PILImage.new("RGB", (100, 100), color=(255, 0, 0))
        img.save(str(fake_img), format="PNG")

        _seed_alert(alert_svc, image_path=str(fake_img))

        meta = report_svc.generate(
            ReportRequest(format=ReportFormat.PDF, include_screenshots=True)
        )
        path = tmp_path / "reports" / meta.file_path
        assert path.exists()
        assert meta.file_size_bytes > 500


# ===========================================================================
# Report Listing & Retrieval Tests
# ===========================================================================
class TestReportListing:
    def test_list_reports(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_alert(alert_svc)

        report_svc.generate(ReportRequest(format=ReportFormat.CSV))
        report_svc.generate(ReportRequest(format=ReportFormat.PDF))

        reports = report_svc.list_reports()
        assert len(reports) == 2

    def test_list_reports_sorted_newest_first(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_alert(alert_svc)

        meta1 = report_svc.generate(ReportRequest(format=ReportFormat.CSV))
        meta2 = report_svc.generate(ReportRequest(format=ReportFormat.PDF))

        reports = report_svc.list_reports()
        # Most recent should be first
        assert reports[0].report_id == meta2.report_id

    def test_get_report_path(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_alert(alert_svc)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.CSV))
        path = report_svc.get_report_path(meta.report_id)

        assert path.exists()
        assert path.name == meta.filename

    def test_get_nonexistent_report(self, tmp_path):
        _, _, report_svc = _make_services(tmp_path)

        with pytest.raises(ReportNotFoundError):
            report_svc.get_report_path("RPT-nonexistent")

    def test_report_metadata_fields(self, tmp_path):
        alert_svc, _, report_svc = _make_services(tmp_path)
        _seed_alert(alert_svc)

        meta = report_svc.generate(ReportRequest(format=ReportFormat.EXCEL))

        assert meta.report_id.startswith("RPT-")
        assert meta.format == ReportFormat.EXCEL
        assert meta.filename.endswith(".xlsx")
        assert meta.generated_at is not None
        assert meta.file_size_bytes > 0
