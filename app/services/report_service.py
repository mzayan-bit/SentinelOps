"""
SentinelOps — Report Service
================================
Generates violation reports in CSV, Excel, and PDF formats.

Each report includes:
- A full violation table (alert records)
- Summary statistics (totals by type, severity, status, compliance)
- Charts rendered via ``ChartBuilder`` (Excel / PDF only)
- Embedded screenshots from alert ``image_path`` fields (PDF only)

Reports are persisted to ``artifacts/reports/``.

Usage::

    from app.services.report_service import ReportService

    report_svc = ReportService(alert_service, analytics_service)
    meta = report_svc.generate(ReportRequest(format=ReportFormat.PDF))
"""

from __future__ import annotations

import csv
import io
import json
import logging
import shutil
import tempfile
import uuid
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from app.models.alert import AlertFilter
from app.models.report import ReportFormat, ReportMetadata, ReportRequest
from app.services.alert_service import AlertService
from app.services.analytics_service import AnalyticsService
from app.services.chart_builder import ChartBuilder

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logger = logging.getLogger("sentinelops.report_service")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEFAULT_REPORTS_DIR = Path("artifacts/reports")
_INDEX_FILE = "reports_index.json"
_MAX_PDF_VIOLATIONS = 100
_ALERT_COLUMNS = [
    "alert_id",
    "timestamp",
    "camera_id",
    "alert_type",
    "severity",
    "status",
    "confidence",
    "assigned_to",
    "notes",
    "image_path",
]


class ReportNotFoundError(Exception):
    """Raised when a requested report does not exist."""


class ReportService:
    """Generates and manages violation reports.

    Parameters
    ----------
    alert_service : AlertService
        Source of raw alert data.
    analytics_service : AnalyticsService
        Source of aggregated metrics for charts/statistics.
    reports_dir : Path
        Output directory for generated reports.
    """

    def __init__(
        self,
        alert_service: AlertService,
        analytics_service: AnalyticsService,
        reports_dir: Path = DEFAULT_REPORTS_DIR,
    ) -> None:
        self._alert_service = alert_service
        self._analytics_service = analytics_service
        self._dir = reports_dir
        self._dir.mkdir(parents=True, exist_ok=True)
        self._index_path = self._dir / _INDEX_FILE
        self._index: dict[str, dict[str, Any]] = self._load_index()

    # ------------------------------------------------------------------
    # Index persistence
    # ------------------------------------------------------------------
    def _load_index(self) -> dict[str, dict[str, Any]]:
        if self._index_path.exists():
            return json.loads(self._index_path.read_text(encoding="utf-8"))
        return {}

    def _save_index(self) -> None:
        self._index_path.write_text(
            json.dumps(self._index, indent=2, ensure_ascii=False, default=str),
            encoding="utf-8",
        )

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def generate(self, request: ReportRequest) -> ReportMetadata:
        """Generate a report and return its metadata."""
        report_id = self._generate_id()
        ts = datetime.now(timezone.utc)
        ext_map = {
            ReportFormat.CSV: "csv",
            ReportFormat.EXCEL: "xlsx",
            ReportFormat.PDF: "pdf",
        }
        ext = ext_map[request.format]
        filename = (
            f"sentinelops_report_{ts.strftime('%Y%m%d_%H%M%S')}"
            f"_{report_id[-8:]}.{ext}"
        )

        # Fetch alert data
        alerts = self._fetch_alerts(request.date_from, request.date_to)

        # Dispatch to format-specific generator
        generator = {
            ReportFormat.CSV: self._generate_csv,
            ReportFormat.EXCEL: self._generate_excel,
            ReportFormat.PDF: self._generate_pdf,
        }[request.format]

        file_path = generator(
            alerts=alerts,
            request=request,
            filename=filename,
        )

        size = file_path.stat().st_size
        meta = ReportMetadata(
            report_id=report_id,
            format=request.format,
            filename=filename,
            generated_at=ts,
            file_path=str(file_path.relative_to(self._dir)),
            file_size_bytes=size,
        )

        # Persist to index
        self._index[report_id] = meta.model_dump(mode="json")
        self._save_index()

        logger.info(
            "Report generated: %s (%s, %d bytes)",
            report_id, request.format.value, size,
        )
        return meta

    def list_reports(self) -> list[ReportMetadata]:
        """Return metadata for all generated reports, newest first."""
        reports = [ReportMetadata(**v) for v in self._index.values()]
        reports.sort(key=lambda r: r.generated_at, reverse=True)
        return reports

    def get_report_path(self, report_id: str) -> Path:
        """Resolve a report ID to its absolute file path.

        Raises
        ------
        ReportNotFoundError
        """
        entry = self._index.get(report_id)
        if entry is None:
            raise ReportNotFoundError(f"Report '{report_id}' not found.")
        path = self._dir / entry["file_path"]
        if not path.exists():
            raise ReportNotFoundError(f"Report file missing: {path}")
        return path

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _generate_id() -> str:
        ts = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
        return f"RPT-{ts}-{uuid.uuid4().hex[:8]}"

    def _fetch_alerts(
        self,
        date_from: datetime | None = None,
        date_to: datetime | None = None,
    ) -> list[dict[str, Any]]:
        filters = AlertFilter(date_from=date_from, date_to=date_to)
        alerts = self._alert_service.list_alerts(filters)
        return [a.to_dict() for a in alerts]

    def _build_stats(self, alerts: list[dict[str, Any]]) -> dict[str, Any]:
        """Compute summary statistics from raw alert dicts."""
        total = len(alerts)
        by_type = dict(Counter(a.get("alert_type", "Unknown") for a in alerts))
        by_severity = dict(Counter(a.get("severity", "Unknown") for a in alerts))
        by_status = dict(Counter(a.get("status", "Unknown") for a in alerts))

        ppe_types = {"No Helmet", "No Vest"}
        ppe_violations = sum(1 for a in alerts if a.get("alert_type") in ppe_types)
        compliance_rate = round((total - ppe_violations) / total, 4) if total else 1.0

        return {
            "total_violations": total,
            "by_type": by_type,
            "by_severity": by_severity,
            "by_status": by_status,
            "ppe_violations": ppe_violations,
            "compliance_rate": compliance_rate,
        }

    def _build_charts(
        self, request: ReportRequest
    ) -> dict[str, Path]:
        """Render all analytics charts and return name→path mapping."""
        chart_dir = Path(tempfile.mkdtemp(prefix="sentinelops_charts_"))
        builder = ChartBuilder(output_dir=chart_dir)

        vpd = self._analytics_service.violations_per_day(
            request.date_from, request.date_to
        )
        vpc = self._analytics_service.violations_per_camera(
            request.date_from, request.date_to
        )
        cr = self._analytics_service.compliance_rate(
            request.date_from, request.date_to
        )
        ht = self._analytics_service.hourly_trends(
            request.date_from, request.date_to
        )
        tv = self._analytics_service.top_violation_types(
            request.date_from, request.date_to
        )

        return {
            "violations_per_day": builder.violations_per_day_chart(vpd),
            "violations_per_camera": builder.violations_per_camera_chart(vpc),
            "compliance_rate": builder.compliance_pie_chart(cr),
            "hourly_trends": builder.hourly_trends_chart(ht),
            "top_violations": builder.top_violations_chart(tv),
        }

    # ------------------------------------------------------------------
    # CSV generator
    # ------------------------------------------------------------------
    def _generate_csv(
        self,
        alerts: list[dict[str, Any]],
        request: ReportRequest,
        filename: str,
    ) -> Path:
        """Generate a CSV report with alert table + summary statistics."""
        out_path = self._dir / filename

        with open(out_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            # Header
            writer.writerow(_ALERT_COLUMNS)

            # Data rows
            for alert in alerts:
                writer.writerow(
                    [alert.get(col, "") for col in _ALERT_COLUMNS]
                )

            # Blank separator
            writer.writerow([])
            writer.writerow(["--- SUMMARY STATISTICS ---"])

            # Statistics
            stats = self._build_stats(alerts)
            writer.writerow(["Total Violations", stats["total_violations"]])
            writer.writerow(
                ["PPE Compliance Rate", f"{stats['compliance_rate']:.2%}"]
            )
            writer.writerow(["PPE Violations", stats["ppe_violations"]])
            writer.writerow([])

            writer.writerow(["Violations by Type"])
            for vtype, count in sorted(
                stats["by_type"].items(), key=lambda x: x[1], reverse=True
            ):
                writer.writerow(["", vtype, count])

            writer.writerow([])
            writer.writerow(["Violations by Severity"])
            for sev, count in sorted(
                stats["by_severity"].items(), key=lambda x: x[1], reverse=True
            ):
                writer.writerow(["", sev, count])

            writer.writerow([])
            writer.writerow(["Violations by Status"])
            for status, count in sorted(
                stats["by_status"].items(), key=lambda x: x[1], reverse=True
            ):
                writer.writerow(["", status, count])

        logger.debug("CSV report generated: %s", out_path)
        return out_path

    # ------------------------------------------------------------------
    # Excel generator
    # ------------------------------------------------------------------
    def _generate_excel(
        self,
        alerts: list[dict[str, Any]],
        request: ReportRequest,
        filename: str,
    ) -> Path:
        """Generate a multi-sheet Excel workbook."""
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        out_path = self._dir / filename
        wb = Workbook()

        # -- Sheet 1: Violations ----------------------------------------
        ws_viol = wb.active
        ws_viol.title = "Violations"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")

        # Header row
        for col_idx, col_name in enumerate(_ALERT_COLUMNS, 1):
            cell = ws_viol.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, alert in enumerate(alerts, 2):
            for col_idx, col_name in enumerate(_ALERT_COLUMNS, 1):
                ws_viol.cell(row=row_idx, column=col_idx, value=alert.get(col_name, ""))

        # Auto-width columns
        for col_idx in range(1, len(_ALERT_COLUMNS) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(str(ws_viol.cell(row=r, column=col_idx).value or ""))
                for r in range(1, min(len(alerts) + 2, 52))  # sample first 50 rows
            )
            ws_viol.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Auto-filter
        if alerts:
            ws_viol.auto_filter.ref = (
                f"A1:{get_column_letter(len(_ALERT_COLUMNS))}{len(alerts) + 1}"
            )

        # -- Sheet 2: Statistics ----------------------------------------
        ws_stats = wb.create_sheet("Statistics")
        stats = self._build_stats(alerts)

        stat_rows = [
            ["SUMMARY STATISTICS"],
            [],
            ["Total Violations", stats["total_violations"]],
            ["PPE Compliance Rate", f"{stats['compliance_rate']:.2%}"],
            ["PPE Violations", stats["ppe_violations"]],
            [],
            ["VIOLATIONS BY TYPE"],
        ]
        for vtype, count in sorted(
            stats["by_type"].items(), key=lambda x: x[1], reverse=True
        ):
            stat_rows.append([vtype, count])

        stat_rows.extend([[], ["VIOLATIONS BY SEVERITY"]])
        for sev, count in sorted(
            stats["by_severity"].items(), key=lambda x: x[1], reverse=True
        ):
            stat_rows.append([sev, count])

        stat_rows.extend([[], ["VIOLATIONS BY STATUS"]])
        for status, count in sorted(
            stats["by_status"].items(), key=lambda x: x[1], reverse=True
        ):
            stat_rows.append([status, count])

        for row_idx, row_data in enumerate(stat_rows, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
                if row_idx in (1, 7) or (
                    row_data and isinstance(row_data[0], str) and row_data[0].startswith("VIOLATIONS BY")
                ):
                    cell.font = Font(bold=True, size=12)

        ws_stats.column_dimensions["A"].width = 30
        ws_stats.column_dimensions["B"].width = 20

        # -- Sheet 3: Charts -------------------------------------------
        charts = None
        if request.include_charts:
            ws_charts = wb.create_sheet("Charts")
            charts = self._build_charts(request)

            current_row = 1
            for chart_name, chart_path in charts.items():
                if chart_path.exists():
                    ws_charts.cell(
                        row=current_row, column=1,
                        value=chart_name.replace("_", " ").title(),
                    ).font = Font(bold=True, size=12)
                    current_row += 1
                    img = XlImage(str(chart_path))
                    img.width = 750
                    img.height = 375
                    ws_charts.add_image(img, f"A{current_row}")
                    current_row += 22  # space for image

        wb.save(out_path)

        # Clean up temp chart files (must be AFTER wb.save because
        # openpyxl reads image files lazily during serialisation)
        if request.include_charts and charts:
            chart_dir = next(iter(charts.values())).parent
            shutil.rmtree(chart_dir, ignore_errors=True)

        logger.debug("Excel report generated: %s", out_path)
        return out_path

    # ------------------------------------------------------------------
    # PDF generator
    # ------------------------------------------------------------------
    def _generate_pdf(
        self,
        alerts: list[dict[str, Any]],
        request: ReportRequest,
        filename: str,
    ) -> Path:
        """Generate a structured PDF report."""
        from fpdf import FPDF

        out_path = self._dir / filename
        pdf = _SentinelOpsPDF(title=request.title)
        pdf.set_auto_page_break(auto=True, margin=20)

        # -- Title page -------------------------------------------------
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 24)
        pdf.ln(30)
        pdf.cell(0, 15, request.title, new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.set_font("Helvetica", "", 12)

        date_range_text = "All Time"
        if request.date_from and request.date_to:
            date_range_text = (
                f"{request.date_from.strftime('%Y-%m-%d')} to "
                f"{request.date_to.strftime('%Y-%m-%d')}"
            )
        elif request.date_from:
            date_range_text = f"From {request.date_from.strftime('%Y-%m-%d')}"
        elif request.date_to:
            date_range_text = f"Until {request.date_to.strftime('%Y-%m-%d')}"

        pdf.cell(0, 10, f"Date Range: {date_range_text}", new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.cell(
            0, 10,
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            new_x="LMARGIN", new_y="NEXT", align="C",
        )
        pdf.ln(10)

        # -- Executive summary ------------------------------------------
        stats = self._build_stats(alerts)
        pdf.add_page()
        pdf.chapter_title("Executive Summary")

        summary_data = [
            ["Metric", "Value"],
            ["Total Violations", str(stats["total_violations"])],
            ["PPE Compliance Rate", f"{stats['compliance_rate']:.2%}"],
            ["PPE Violations", str(stats["ppe_violations"])],
        ]
        pdf.data_table(summary_data, col_widths=[95, 95])
        pdf.ln(8)

        # By type
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, "Violations by Type", new_x="LMARGIN", new_y="NEXT")
        type_data = [["Type", "Count"]]
        for vtype, count in sorted(
            stats["by_type"].items(), key=lambda x: x[1], reverse=True
        ):
            type_data.append([vtype, str(count)])
        pdf.data_table(type_data, col_widths=[130, 60])
        pdf.ln(8)

        # By severity
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, "Violations by Severity", new_x="LMARGIN", new_y="NEXT")
        sev_data = [["Severity", "Count"]]
        for sev, count in sorted(
            stats["by_severity"].items(), key=lambda x: x[1], reverse=True
        ):
            sev_data.append([sev, str(count)])
        pdf.data_table(sev_data, col_widths=[130, 60])

        # -- Charts -----------------------------------------------------
        if request.include_charts:
            charts = self._build_charts(request)

            pdf.add_page()
            pdf.chapter_title("Analytics Charts")

            for chart_name, chart_path in charts.items():
                if chart_path.exists():
                    # Check if we need a new page
                    if pdf.get_y() > 160:
                        pdf.add_page()
                    pdf.set_font("Helvetica", "B", 11)
                    pdf.cell(
                        0, 8,
                        chart_name.replace("_", " ").title(),
                        new_x="LMARGIN", new_y="NEXT",
                    )
                    pdf.image(str(chart_path), x=10, w=190)
                    pdf.ln(8)

            # Clean up temp charts
            chart_dir = next(iter(charts.values())).parent
            shutil.rmtree(chart_dir, ignore_errors=True)

        # -- Violation detail table --------------------------------------
        pdf.add_page()
        pdf.chapter_title("Violation Details")

        display_alerts = alerts[:_MAX_PDF_VIOLATIONS]
        detail_cols = ["alert_id", "timestamp", "camera_id", "alert_type", "severity", "status"]
        col_widths = [38, 35, 22, 32, 22, 22]

        # Header
        header_row = [c.replace("_", " ").title() for c in detail_cols]
        detail_data = [header_row]
        for alert in display_alerts:
            row = []
            for col in detail_cols:
                val = str(alert.get(col, ""))
                # Truncate long values
                if len(val) > 22 and col != "alert_id":
                    val = val[:19] + "..."
                row.append(val)
            detail_data.append(row)

        pdf.data_table(detail_data, col_widths=col_widths)

        if len(alerts) > _MAX_PDF_VIOLATIONS:
            pdf.ln(5)
            pdf.set_font("Helvetica", "I", 9)
            pdf.cell(
                0, 6,
                f"Showing {_MAX_PDF_VIOLATIONS} of {len(alerts)} violations. "
                "Export to CSV or Excel for the full dataset.",
                new_x="LMARGIN", new_y="NEXT",
            )

        # -- Screenshots -------------------------------------------------
        if request.include_screenshots:
            screenshots = [
                a for a in alerts
                if a.get("image_path") and Path(a["image_path"]).exists()
            ]
            if screenshots:
                pdf.add_page()
                pdf.chapter_title("Evidence Screenshots")

                for alert in screenshots[:20]:  # limit to 20 screenshots
                    img_path = alert["image_path"]
                    if pdf.get_y() > 180:
                        pdf.add_page()
                    pdf.set_font("Helvetica", "B", 9)
                    pdf.cell(
                        0, 6,
                        f"{alert.get('alert_id', 'N/A')} - "
                        f"{alert.get('alert_type', 'Unknown')} - "
                        f"{alert.get('camera_id', 'Unknown')}",
                        new_x="LMARGIN", new_y="NEXT",
                    )
                    try:
                        pdf.image(img_path, x=10, w=120)
                    except Exception:
                        pdf.set_font("Helvetica", "I", 9)
                        pdf.cell(
                            0, 6, "[Screenshot could not be loaded]",
                            new_x="LMARGIN", new_y="NEXT",
                        )
                    pdf.ln(6)

        pdf.output(str(out_path))
        logger.debug("PDF report generated: %s", out_path)
        return out_path


# ---------------------------------------------------------------------------
# Custom FPDF subclass for SentinelOps branding
# ---------------------------------------------------------------------------
class _SentinelOpsPDF:
    """Thin wrapper around FPDF with SentinelOps header/footer branding."""

    def __new__(cls, title: str = "SentinelOps Report"):
        from fpdf import FPDF

        class _PDF(FPDF):
            _report_title = title

            def header(self):
                self.set_font("Helvetica", "B", 9)
                self.set_text_color(120, 120, 120)
                self.cell(0, 8, f"SentinelOps  |  {self._report_title}", new_x="LMARGIN", new_y="NEXT")
                self.set_draw_color(200, 200, 200)
                self.line(10, self.get_y(), 200, self.get_y())
                self.ln(4)

            def footer(self):
                self.set_y(-15)
                self.set_font("Helvetica", "I", 8)
                self.set_text_color(150, 150, 150)
                self.cell(0, 10, f"Page {self.page_no()}/{{nb}}", align="C")

            def chapter_title(self, title: str):
                self.set_font("Helvetica", "B", 16)
                self.set_text_color(26, 26, 46)
                self.cell(0, 12, title, new_x="LMARGIN", new_y="NEXT")
                self.set_draw_color(15, 155, 142)
                self.set_line_width(0.8)
                self.line(10, self.get_y(), 80, self.get_y())
                self.set_line_width(0.2)
                self.ln(6)
                self.set_text_color(0, 0, 0)

            def data_table(
                self,
                data: list[list[str]],
                col_widths: list[int] | None = None,
            ):
                """Render a table with header row styling."""
                if not data:
                    return
                if col_widths is None:
                    n_cols = len(data[0])
                    col_widths = [190 // n_cols] * n_cols

                # Header row
                self.set_font("Helvetica", "B", 9)
                self.set_fill_color(26, 26, 46)
                self.set_text_color(255, 255, 255)
                for j, cell_val in enumerate(data[0]):
                    self.cell(col_widths[j], 7, str(cell_val), border=1, fill=True, align="C")
                self.ln()

                # Data rows
                self.set_font("Helvetica", "", 8)
                self.set_text_color(30, 30, 30)
                fill = False
                for row in data[1:]:
                    if self.get_y() > 265:
                        self.add_page()
                        # Re-draw header on new page
                        self.set_font("Helvetica", "B", 9)
                        self.set_fill_color(26, 26, 46)
                        self.set_text_color(255, 255, 255)
                        for j, cell_val in enumerate(data[0]):
                            self.cell(col_widths[j], 7, str(cell_val), border=1, fill=True, align="C")
                        self.ln()
                        self.set_font("Helvetica", "", 8)
                        self.set_text_color(30, 30, 30)
                        fill = False
                    if fill:
                        self.set_fill_color(240, 240, 245)
                    else:
                        self.set_fill_color(255, 255, 255)
                    for j, cell_val in enumerate(row):
                        self.cell(col_widths[j], 6, str(cell_val), border=1, fill=True)
                    self.ln()
                    fill = not fill

        pdf = _PDF()
        pdf.alias_nb_pages()
        return pdf
